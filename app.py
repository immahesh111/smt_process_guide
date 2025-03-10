from flask import Flask, render_template, request, redirect, url_for, send_file
from werkzeug.utils import secure_filename
import os
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)

# Define upload folder for images
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure the upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Function to create PDF
def create_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="SMT Process Guide Submission", ln=1, align="C")
    pdf.ln(10)

    for key, value in data.items():
        pdf.cell(200, 7, txt=f"{key}: {str(value)}", ln=1)

    pdf.output("smt_process_guide.pdf")

# Function to create Excel file
def create_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "SMT Process Guide Data"

    # Write the headers
    headers = list(data.keys())
    ws.append(headers)

    # Write the data
    values = list(data.values())
    ws.append(values)

    wb.save("smt_process_guide.xlsx")

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    if request.method == 'POST':
        try:
            # Extract form data including new fields
            date = request.form.get('date')
            shift = request.form.get('shift')
            failure_time = request.form.get('failureTime')
            failure_station = request.form.get('failureStation')
            pcba_array = request.form.get('pcbaArray')
            serial_number = request.form.get('serialNumber')
            side = request.form.get('side')
            spi_date_time = request.form.get('spiDateTime')
            spi_volume = request.form.get('spiVolume')
            spi_height = request.form.get('spiHeight')
            spi_area = request.form.get('spiArea')
            pre_aoi_date_time = request.form.get('preAoiDateTime')
            pre_aoi_call = request.form.get('preAoiCall')
            pre_aoi_issue = request.form.get('preAoiIssue')
            post_aoi_date_time = request.form.get('postAoiDateTime')
            post_aoi_call = request.form.get('postAoiCall')
            post_aoi_issue = request.form.get('postAoiIssue')
            error_code = request.form.get('errorCode')
            failure_symptom = request.form.get('failureSymptom')
            root_cause = request.form.get('rootCause')
            corrective_action = request.form.get('correctiveAction')
            current_yield = request.form.get('currentYield')  # New field
            improved_yield = request.form.get('improvedYield')  # New field
            fa_done_by = request.form.get('faDoneBy')

            xray_image = request.files.get('xrayImage')
            ct_scan_image = request.files.get('ctScanImage')
            spi_image = request.files.get('spiImage')

            # Save images if they exist
            xray_filename = None
            ct_scan_filename = None
            spi_image_filename = None

            if xray_image and xray_image.filename != '':
                xray_filename = secure_filename(xray_image.filename)
                xray_image.save(os.path.join(app.config['UPLOAD_FOLDER'], xray_filename))

            if ct_scan_image and ct_scan_image.filename != '':
                ct_scan_filename = secure_filename(ct_scan_image.filename)
                ct_scan_image.save(os.path.join(app.config['UPLOAD_FOLDER'], ct_scan_filename))

            if spi_image and spi_image.filename != '':
                spi_image_filename = secure_filename(spi_image.filename)
                spi_image.save(os.path.join(app.config['UPLOAD_FOLDER'], spi_image_filename))

            # Prepare data dictionary with new fields
            data = {
                'Date': date,
                'Shift': shift,
                'Failure_Time': failure_time,
                'Failure_Station': failure_station,
                'PCBA_Array': pcba_array,
                'Serial_Number': serial_number,
                'Side': side,
                'SPI_Date_and_Time': spi_date_time,
                'SPI_Volume': spi_volume,
                'SPI_Height': spi_height,
                'SPI_Area': spi_area,
                'Pre_AOI_Date_and_Time': pre_aoi_date_time,
                'Pre_AOI_Call': pre_aoi_call,
                'Pre_AOI_Issue': pre_aoi_issue,
                'Post_AOI_Date_and_Time': post_aoi_date_time,
                'Post_AOI_Call': post_aoi_call,
                'Post_AOI_Issue': post_aoi_issue,
                'Error_Code': error_code,
                'Failure_Symptom': failure_symptom,
                'Root_Cause': root_cause,
                'Corrective_Action': corrective_action,
                'Current_Yield': current_yield,  # New field
                'Improved_Yield': improved_yield,  # New field
                'FA_Done_By': fa_done_by,
                'X_Ray_Image': xray_filename if xray_filename else "No Image Uploaded",
                'CT_Scan_Image': ct_scan_filename if ct_scan_filename else "No Image Uploaded",
                'SPI_Image': spi_image_filename if spi_image_filename else "No Image Uploaded",
            }

            return render_template('response.html', **data, data_str=str(data),
                                 xray_image=xray_filename, 
                                 ct_scan_image=ct_scan_filename, 
                                 spi_image=spi_image_filename)
        except Exception as e:
            return f"An error occurred: {e}"

@app.route('/download_pdf', methods=['POST'])
def download_pdf():
    try:
        data_str = request.form.get('data_str')
        data = eval(data_str)
        create_pdf(data)
        return send_file("smt_process_guide.pdf", as_attachment=True)
    except Exception as e:
        return f"Failed to download PDF: {e}"

@app.route('/download_excel', methods=['POST'])
def download_excel():
    try:
        data_str = request.form.get('data_str')
        data = eval(data_str)
        create_excel(data)
        return send_file("smt_process_guide.xlsx", as_attachment=True)
    except Exception as e:
        return f"Failed to download Excel: {e}"

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=2000)
